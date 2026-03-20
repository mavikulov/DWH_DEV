import os
from pathlib import Path

import pandas as pd
import openpyxl

from utils import (
    find_end_date_by_name,
    find_end_date_in_table,
    convert_and_rename_columns,
    extract_final_date,
    find_header_by_columns,
)


def parse_vladimir_esv(xlsx_file_path):
    HEADER_TABLE_NUM = 14
    wb = openpyxl.load_workbook(filename=xlsx_file_path)
    sheet = wb.active
    df = pd.read_excel(
        xlsx_file_path, sheet_name=wb.sheetnames[0], header=HEADER_TABLE_NUM
    )
    unnamed_columns = [column for column in df.columns if column.startswith("Unnamed")]
    dropped_df = df.drop(columns=unnamed_columns, axis=1)
    dropped_df["Итого"] = pd.to_numeric(df["Итого"].str.strip().str.replace(",", "."))
    end_date = find_end_date_in_table(sheet)
    if end_date is None:
        end_date = find_end_date_by_name(
            os.path.basename(os.path.normpath(xlsx_file_path))
        )

    dropped_df["Дата"] = end_date
    converted_df = convert_and_rename_columns(
        dropped_df[:-1], ["Номер ЛС", "Дата", "Итого"]
    )
    return converted_df


def parse_vladimir_tplus(xlsx_file_path):
    wb = openpyxl.load_workbook(filename=xlsx_file_path)
    sheet = wb.active
    df = pd.read_excel(xlsx_file_path, sheet_name=wb.sheetnames[0])
    dropped_df = df.drop(columns=["ОПиОК"], axis=1)
    end_date = find_end_date_in_table(sheet)
    if end_date is None:
        end_date = find_end_date_by_name(
            os.path.basename(os.path.normpath(xlsx_file_path))
        )
    dropped_df["Дата"] = end_date
    converted_df = convert_and_rename_columns(
        dropped_df, ["ЛС", "Дата", "Сумма оплаты"]
    )
    return converted_df


def parse_vladimir_up_rkc(xlsx_file_path):
    wb = openpyxl.load_workbook(filename=xlsx_file_path)
    sheet = wb.active
    df = pd.read_excel(xlsx_file_path, sheet_name=wb.sheetnames[0])
    end_date = find_end_date_in_table(sheet)
    if end_date is None:
        end_date = find_end_date_by_name(
            os.path.basename(os.path.normpath(xlsx_file_path))
        )

    df["Дата"] = end_date
    converted_df = convert_and_rename_columns(df, ["Лицевой счет", "Дата", "Оплата"])
    return converted_df


def parse_mosobl_eirc(file_path):
    file = Path(file_path)
    file_format = file.suffix
    if file_format == ".csv":
        df = pd.read_csv(file_path, encoding="cp1251", sep=";")
    elif file_format == ".xlsx":
        df = pd.read_excel(file_path)

    df = df.iloc[:-1]
    df_selected = df[
        ["[Номер ЛС]", "[Дата оплаты]", "[ИТОГО(услуги)]", "[ИТОГО(пени)]"]
    ]

    def safe_convert_to_float(column):
        if column.dtype == "object":
            return pd.to_numeric(
                column.astype(str).str.replace(" ", "").str.replace(",", "."),
                errors="coerce",
            )
        return column

    df_selected["[ИТОГО(услуги)]"] = safe_convert_to_float(
        df_selected["[ИТОГО(услуги)]"]
    )
    df_selected["[ИТОГО(пени)]"] = safe_convert_to_float(df_selected["[ИТОГО(пени)]"])
    df_selected["[ИТОГО(услуги)]"] = df_selected["[ИТОГО(услуги)]"].fillna(0.0)
    df_selected["[ИТОГО(пени)]"] = df_selected["[ИТОГО(пени)]"].fillna(0.0)
    df_selected["[Дата оплаты]"] = df_selected["[Дата оплаты]"].apply(
        extract_final_date
    )
    df_selected["Сумма"] = df_selected["[ИТОГО(услуги)]"] + df_selected["[ИТОГО(пени)]"]
    converted_df = convert_and_rename_columns(
        df_selected, ["[Номер ЛС]", "[Дата оплаты]", "Сумма"]
    )
    return converted_df


def parse_mosoble_mosenergo(file_path):
    if file_path.endswith(".xls"):
        engine = "xlrd"
        header = 0
        date_raspred_column_name = "Дата распределения"
        df = pd.read_excel(file_path, engine=engine, header=header)
    else:
        engine = "openpyxl"
        try:
            df = pd.read_excel(file_path, engine=engine, header=14)
            date_raspred_column_name = "Дата БВ"
            if "ЕЛС" not in df.columns:
                raise ValueError("Колонка 'ЕЛС' не найдена при header=14")
        except Exception:
            df = pd.read_excel(file_path, engine=engine, header=15)
            date_raspred_column_name = "Дата БВ"
            if "ЕЛС" not in df.columns:
                raise ValueError(
                    f"Не удалось найти колонку 'ЕЛС' ни при header=14, ни при header=15 в файле {file_path}"
                )

    converted_df = convert_and_rename_columns(
        df, ["ЕЛС", date_raspred_column_name, "Сумма поступивших ДСб руб"]
    )
    return converted_df


def parse_tula_with_fallback(file_path, header_candidates=(13, 14)):
    required_cols = [
        "Платежный код*",
        "Распределено всего(услуги+пени)",
        "Дата оплаты*",
        "Опердень распределения*",
    ]
    for header_row in header_candidates:
        try:
            df = pd.read_excel(file_path, header=header_row)
            if all(col in df.columns for col in required_cols):
                df = df[list(required_cols)].copy()
                df = df[df["Опердень распределения*"].notna()]
                df["Опердень распределения*"] = pd.to_datetime(
                    df["Опердень распределения*"], format="%d.%m.%Y", errors="coerce"
                ).dt.strftime("%d.%m.%Y")

                converted_df = convert_and_rename_columns(
                    df=df,
                    original_columns=[
                        "Платежный код*",
                        "Опердень распределения*",
                        "Распределено всего(услуги+пени)",
                    ],
                )
                return converted_df
        except Exception as e:
            print(f"Exception = {e}")
            continue

    raise ValueError(
        f"Не удалось найти корректный заголовок в файле {file_path} "
        f"ни при header={header_candidates[0]}, ни при header={header_candidates[1]}"
    )


def parse_tula(file_path):
    return parse_tula_with_fallback(file_path, header_candidates=(13, 14))


def parse_garant_invest(file_path):
    df = pd.read_excel(file_path, header=4, engine="xlrd")
    mask_not_nan_df = df[df["Лицевой счет"].notna()]
    filtered_df = mask_not_nan_df[mask_not_nan_df["Лицевой счет"].str.isdigit()]
    filtered_df = filtered_df[filtered_df["Итого"].notna()]
    selected_df = filtered_df[["Лицевой счет", "Дата оплаты", "Итого"]]
    selected_df["Итого"] = (
        selected_df["Итого"]
        .str.replace("\xa0", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.replace(",", ".", regex=False)
    )

    selected_df["Итого"] = pd.to_numeric(selected_df["Итого"])
    converted_df = convert_and_rename_columns(
        selected_df, ["Лицевой счет", "Дата оплаты", "Итого"]
    )
    return converted_df


def parse_yaroslavl_irc(file_path):
    df = pd.read_excel(file_path, header=2)
    df["Дата платежа"] = df["Дата платежа"].dt.strftime("%d.%m.%Y")
    df["Итого"] = (
        df["Сумма платежа, зачтенного на основной платеж"]
        + df["Сумма платежа, зачтенного на пени"]
    )
    converted_df = convert_and_rename_columns(
        df, ["Лицевой счет", "Дата платежа", "Итого"]
    )
    return converted_df


def parse_yaroslavl_tns(file_path, header=0):
    required_cols = ["Сумма оплаты за услугу", "Дата оплаты"]
    header = find_header_by_columns(file_path, required_cols, max_header=6)
    df = pd.read_excel(file_path, header=header)
    try:
        df = df.rename(columns={"Unnamed: 1": "ЛС"})
    except:
        pass

    if not {"Сумма оплаты за услугу", "Дата оплаты"}.issubset(df.columns):
        raise ValueError("Требуемые колонки отсутствуют")

    df["Итого"] = df["Сумма оплаты за услугу"] + df["Сумма оплаты пени"]
    df = df[df["Дата оплаты"].notna()].copy()

    if pd.api.types.is_datetime64_any_dtype(df["Дата оплаты"]):
        df["Дата оплаты"] = df["Дата оплаты"].dt.strftime("%d.%m.%Y")
    else:
        df["Дата оплаты"] = pd.to_datetime(
            df["Дата оплаты"], errors="coerce"
        ).dt.strftime("%d.%m.%Y")

    if "ЛС" not in df.columns:
        raise ValueError("Колонка 'ЛС' не найдена")

    converted_df = convert_and_rename_columns(df, ["ЛС", "Дата оплаты", "Итого"])
    return converted_df
